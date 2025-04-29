import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from config.settings import EXCEL_HEADERS, HISTORICAL_EXCEL_PATH

def initialize_excel_file(file_path):
    """Initialize an Excel file with headers if it doesn't exist"""
    if not Path(file_path).exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "EOBR Data"
        ws.append(EXCEL_HEADERS)
        wb.save(file_path)

def load_historical_duplicates():
    """Load historical duplicates and control numbers from Excel"""
    historical_duplicates = {}
    max_control_numbers = {}
    
    if Path(HISTORICAL_EXCEL_PATH).exists():
        wb = load_workbook(HISTORICAL_EXCEL_PATH, read_only=True)
        ws = wb.active
        rows = list(ws.rows)[1:]  # Skip header
        
        for row in rows:
            full_dup_key = row[2].value if len(row) > 2 else None
            eobr_number_value = row[4].value if len(row) > 4 else None
            description = row[11].value if len(row) > 11 else None
            
            if full_dup_key and '|' in full_dup_key:
                historical_key = full_dup_key
            else:
                control_number = None
                if eobr_number_value and '-' in eobr_number_value:
                    control_number = eobr_number_value.split('-')[0]
                if control_number and description:
                    cpt_part = description.split(',')[0].strip()
                    historical_key = f"{control_number}|{cpt_part}"
                else:
                    historical_key = full_dup_key or "Unknown"
                    
            if historical_key:
                historical_duplicates[historical_key] = True
                
            if eobr_number_value and '-' in eobr_number_value:
                parts = eobr_number_value.split('-')
                control_number = parts[0]
                try:
                    serial_number = int(parts[1])
                except (ValueError, IndexError):
                    serial_number = 0
                if control_number:
                    max_control_numbers[control_number] = max(
                        max_control_numbers.get(control_number, 0),
                        serial_number
                    )
        wb.close()
        
    return historical_duplicates, max_control_numbers

def append_to_excel(file_path, data):
    """Append data to Excel file"""
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([
        data.get("Release Payment"), data.get("Duplicate Check"), data.get("Full Duplicate Key"),
        data.get("Input File"), data.get("EOBR Number"), data.get("Vendor"), data.get("Mailing Address"),
        data.get("Terms"), data.get("Bill Date"), data.get("Due Date"), data.get("Category"), data.get("Description"),
        data.get("Amount"), data.get("Memo"), data.get("Total"),
    ])
    wb.save(file_path)